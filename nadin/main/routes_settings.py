import io

from flask import Response, current_app, flash, redirect, render_template, request, url_for
from flask_login import current_user, login_required
from openpyxl import Workbook
from sqlalchemy import or_

from nadin.admin.forms import SelectHubForm
from nadin.extensions import db
from nadin.main.forms import UserRolesForm, UserSettingsForm
from nadin.main.routes import bp
from nadin.models.hub import User, UserRoles, Vendor
from nadin.models.order import Order, OrderApproval, OrderCategory, OrderStatus
from nadin.models.project import Project
from nadin.utils import flash_errors, role_forbidden, role_required

################################################################################
# Settings page
################################################################################


@bp.route("/settings/", methods=["GET"])
@login_required
@role_forbidden([UserRoles.default, UserRoles.vendor])
def show_settings():

    if current_user.role == UserRoles.admin:

        search_key = request.args.get("q", type=str)
        page = request.args.get("page", type=int, default=1)

        user_form = UserRolesForm()

        if search_key:
            users, total = User.search(search_key, page, current_app.config["MAX_PER_PAGE"])
        else:
            users = User.query

        users = users.filter(or_(User.role == UserRoles.default, User.hub_id == current_user.hub_id))

        if search_key:
            users = db.paginate(users, page=1, max_per_page=current_app.config["MAX_PER_PAGE"])
            users.total = total
            users.page = page
        else:
            users = users.order_by(User.name, User.email)
            users = db.paginate(users, max_per_page=current_app.config["MAX_PER_PAGE"])

    else:
        users = []
        user_form = UserSettingsForm()
        if current_user.role == UserRoles.initiative:
            current_user.set_initiative_project()

            db.session.commit()

            project = current_user.projects[0]
            user_form.project_name.data = project.name
            user_form.phone.data = project.phone
            user_form.tin.data = project.tin
            user_form.legal_address.data = project.legal_address
            user_form.shipping_address.data = project.shipping_address

    user_form.about_user.full_name.data = current_user.name
    user_form.about_user.email_new.data = current_user.email_new
    user_form.about_user.email_modified.data = current_user.email_modified
    user_form.about_user.email_disapproved.data = current_user.email_disapproved
    user_form.about_user.email_approved.data = current_user.email_approved
    user_form.about_user.email_comment.data = current_user.email_comment
    user_form.about_user.projects.choices = [(p.id, p.name) for p in current_user.projects]

    forms = {
        "select_hub": SelectHubForm(),
    }

    forms["select_hub"].hub_id.choices = [
        (hub.id, f"{hub.id}: {hub.name} ({hub.email})") for hub in Vendor.query.filter(Vendor.hub_id.is_(None)).all()
    ]
    forms["select_hub"].hub_id.default = current_user.hub_id
    forms["select_hub"].process()

    return render_template("main/settings/settings.html", user_form=user_form, users=users, forms=forms)


@bp.route("/settings/", methods=["POST"])
@login_required
@role_forbidden([UserRoles.default, UserRoles.vendor])
def save_settings():

    if current_user.role == UserRoles.admin:
        user_form = UserRolesForm()
    else:
        user_form = UserSettingsForm()

    if current_user.role == UserRoles.initiative:
        projects = current_user.projects
        current_user.set_initiative_project()
        db.session.commit()
    else:
        projects = Project.query
        if current_user.role != UserRoles.admin:
            projects = projects.filter_by(enabled=True)
        projects = projects.filter_by(hub_id=current_user.hub_id)
        projects = projects.order_by(Project.name).all()

    user_form.about_user.projects.choices = [(p.id, p.name) for p in projects]

    if user_form.validate_on_submit():
        if current_user.role == UserRoles.admin:
            user = User.query.filter(User.id == user_form.user_id.data).first()
            if user is None:
                flash("Пользователь не найден.")
                return redirect(url_for("main.show_settings"))
            user.hub_id = current_user.hub_id
            user.role = user_form.role.data
            user.set_initiative_project()
        else:
            user = current_user

        if user.role != UserRoles.initiative:
            if user_form.about_user.projects.data:
                user.projects = Project.query.filter(Project.id.in_(user_form.about_user.projects.data)).all()
            else:
                user.projects = []

        user.email_new = user_form.about_user.email_new.data
        user.email_modified = user_form.about_user.email_modified.data
        user.email_disapproved = user_form.about_user.email_disapproved.data
        user.email_approved = user_form.about_user.email_approved.data
        user.email_comment = user_form.about_user.email_comment.data
        user.name = user_form.about_user.full_name.data.strip()

        if user.role == UserRoles.initiative and isinstance(user_form, UserSettingsForm):
            project_name = user_form.project_name.data.strip() if user_form.project_name.data else current_user.name
            user.projects[0].name = project_name
            user.projects[0].phone = user_form.phone.data.strip()
            user.projects[0].tin = user_form.tin.data.strip()
            user.projects[0].legal_address = user_form.legal_address.data.strip()
            user.projects[0].shipping_address = user_form.shipping_address.data.strip()
            user.projects[0].contact = user.name

        db.session.commit()

        flash("Данные успешно сохранены.")
    else:
        flash_errors(user_form)
        flash_errors(user_form.about_user)

    return redirect(url_for("main.show_settings"))


@bp.route("/users/remove/<int:user_id>")
@login_required
@role_required([UserRoles.admin])
def RemoveUser(user_id):
    user = User.query.filter(
        User.id == user_id,
        or_(User.role == UserRoles.default, User.hub_id == current_user.hub_id),
    ).first()
    if user is None:
        flash("Пользователь не найден.")
        return redirect(url_for("main.show_settings"))

    for order in user.orders:
        order.initiative_id = current_user.id
    db.session.commit()

    db.session.delete(user)
    db.session.commit()

    flash("Пользователь успешно удалён.")
    return redirect(url_for("main.show_settings"))


@bp.route("/users/download")
@login_required
@role_required([UserRoles.admin])
def DownloadUsers():
    users = (
        User.query.filter(or_(User.role == UserRoles.default, User.hub_id == current_user.hub_id))
        .order_by(User.name, User.email)
        .all()
    )
    wb = Workbook()
    ws = wb.active
    for i, header in enumerate(
        [
            "Имя",
            "Телефон",
            "Email",
            "Роль",
            "Адрес",
            "Права",
            "Заметка",
            "Активность",
            "Регистрация",
            "Согласованных заявок пользователя",
            "Сумма согласованных заявок пользователя",
            "Согласовал заявок",
            "Должен согласовать заявок",
            "Номер для согласования",
            "День рождения",
            "Ссылка на дашборд",
        ],
        start=1,
    ):
        ws.cell(1, i).value = header

    for i, user in enumerate(users, start=2):
        ws.cell(i, 1).value = user.name
        ws.cell(i, 3).value = user.email
        ws.cell(i, 6).value = user.role
        ws.cell(i, 8).value = user.last_seen
        ws.cell(i, 9).value = user.registered

        # Orders which user is initiative for
        orders = Order.query.filter_by(initiative_id=user.id).all()
        ws.cell(i, 10).value = len(orders)
        ws.cell(i, 11).value = sum(o.total for o in orders)

        if user.role in [UserRoles.purchaser, UserRoles.validator]:
            # Orders approved by user
            orders = (
                Order.query.filter_by(hub_id=current_user.hub_id)
                .join(OrderApproval)
                .filter_by(user_id=user.id, product_id=None)
                .all()
            )
            ws.cell(i, 12).value = len(orders)

            # Orders to be approved
            orders = Order.query.filter(
                Order.hub_id == current_user.hub_id,
                or_(
                    Order.status == OrderStatus.new,
                    Order.status == OrderStatus.unpayed,
                ),
                ~Order.user_approvals.any(OrderApproval.user_id == user.id),
                ~Order.children.any(),
            )
            orders = orders.join(OrderCategory)
            orders = orders.filter(OrderCategory.category_id.in_([cat.id for cat in user.categories]))
            orders = orders.join(Project)
            orders = orders.filter(Project.id.in_([p.id for p in user.projects]))
            orders = orders.all()
            ws.cell(i, 13).value = len(orders)
            ws.cell(i, 14).value = ", ".join([o.number for o in orders])
        else:
            ws.cell(i, 12).value = 0
            ws.cell(i, 13).value = 0
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return Response(
        buffer,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment;filename=users.xlsx"},
    )
