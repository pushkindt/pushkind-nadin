from datetime import datetime

import pandas as pd
from flask import current_app, flash, redirect, render_template, request, url_for
from flask_login import current_user, login_required
from openpyxl import load_workbook

from nadin.extensions import db
from nadin.main.forms import AddProjectForm, EditProjectForm, UploadProjectsForm
from nadin.main.routes import bp
from nadin.models.hub import UserRoles
from nadin.models.project import Project, ProjectOrderHistory, ProjectPriceLevel
from nadin.utils import flash_errors, role_forbidden


def projects_excel_to_df(excel_file) -> tuple[pd.DataFrame, pd.DataFrame]:
    wb = load_workbook(excel_file, read_only=True, data_only=True)
    sheet = wb.active
    start_row = 0
    for i, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        if "Общая информация" in row:
            start_row = i
            break
    header_row = start_row + 1
    header = tuple(sheet.iter_rows(min_row=header_row, max_row=header_row, values_only=True))[0]
    data_row = header_row + 2
    header_indexes = {
        "uid": header.index("Код"),
        "name": header.index("Название"),
        "tin": header.index("ИНН"),
        "phone": header.index("Телефон"),
        "email": header.index("Мейл"),
        "contact": header.index("Контактное лицо"),
        "note": header.index("Комментарий"),
        "legal_address": header.index("Адрес (сцеп)"),
        "shipping_address": header.index("Адрес доставки"),
        "price_level": header.index("Канал"),
        "last_order_date": 19,
        "2019": header.index(2019),
        "2020": header.index(2020),
        "2021": header.index(2021),
        "2022": header.index(2022),
        "2023": header.index(2023),
        "2024": header.index(2024),
        "active_month_2023": header.index("Прошлый год"),
        "active_month_2024": header.index("Текущий год (2024)"),
    }

    projects = []
    project_order_history = {}
    for row in sheet.iter_rows(min_row=data_row, values_only=True):
        uid = str(row[header_indexes["uid"]]).strip() if row[header_indexes["uid"]] else None
        name = str(row[header_indexes["name"]]).strip() if row[header_indexes["name"]] else None
        if not name or not uid:
            continue
        project = {
            "uid": uid,
            "name": name,
            "tin": str(row[header_indexes["tin"]]) if row[header_indexes["tin"]] else None,
            "phone": str(row[header_indexes["phone"]]) if row[header_indexes["phone"]] else None,
            "email": str(row[header_indexes["email"]]) if row[header_indexes["email"]] else None,
            "contact": str(row[header_indexes["contact"]]) if row[header_indexes["contact"]] else None,
            "note": str(row[header_indexes["note"]]) if row[header_indexes["note"]] else None,
            "legal_address": (
                str(row[header_indexes["legal_address"]]) if row[header_indexes["legal_address"]] else None
            ),
            "shipping_address": (
                str(row[header_indexes["shipping_address"]]) if row[header_indexes["shipping_address"]] else None
            ),
            "last_order_date": (
                datetime.strptime(str(row[header_indexes["last_order_date"]]), "%m.%Y").date().isoformat()
                if row[header_indexes["last_order_date"]]
                else None
            ),
            "price_level": ProjectPriceLevel.from_pretty(str(row[header_indexes["price_level"]]).strip()).name,
        }
        project_order_history[project["uid"]] = {
            "year": [2019, 2020, 2021, 2022, 2023, 2024],
            "total": [
                row[header_indexes["2019"]] if row[header_indexes["2019"]] else 0,
                row[header_indexes["2020"]] if row[header_indexes["2020"]] else 0,
                row[header_indexes["2021"]] if row[header_indexes["2021"]] else 0,
                row[header_indexes["2022"]] if row[header_indexes["2022"]] else 0,
                row[header_indexes["2023"]] if row[header_indexes["2023"]] else 0,
                row[header_indexes["2024"]] if row[header_indexes["2024"]] else 0,
            ],
            "active_months": [
                12,
                12,
                12,
                12,
                row[header_indexes["active_month_2023"]] if row[header_indexes["active_month_2023"]] else 0,
                row[header_indexes["active_month_2024"]] if row[header_indexes["active_month_2024"]] else 0,
            ],
        }
        projects.append(project)
    projects = pd.DataFrame(projects)
    projects_order_history = pd.DataFrame(project_order_history)
    projects_order_history = (
        projects_order_history.transpose()
        .explode(["year", "total", "active_months"])
        .reset_index()
        .rename(columns={"index": "uid"})
    )
    return projects, projects_order_history


@bp.route("/projects/", methods=["GET"])
@bp.route("/projects/show", methods=["GET"])
@login_required
@role_forbidden([UserRoles.default, UserRoles.vendor, UserRoles.initiative])
def ShowProjects():

    show_add_project = request.args.get("add_project", default=False, type=bool)
    search_key = request.args.get("q", type=str)
    search_fields = request.args.getlist("field", type=str)
    page = request.args.get("page", type=int, default=1)
    price_level = request.args.get("price_level", type=str)
    last_order_date = request.args.get("last_order_date", type=int)

    if price_level in ProjectPriceLevel.__members__:
        price_level = ProjectPriceLevel[price_level]
    else:
        price_level = None

    if last_order_date not in [1, 6, 12]:
        last_order_date = None

    fast_search = [
        "Москва",
        "Санкт-Петербург",
        "Саратов",
        "Пермь",
        "Уфа",
        "Воронеж",
        "Новосибирск",
        "Люберцы",
        "Челябинск",
        "Владивосток",
        "Казань",
        "Якутск",
        "Томск",
        "Нижний Новгород",
        "Екатеринбург",
        "Тверь",
    ]

    if search_key:
        projects, total = Project.search(search_key, page, current_app.config["MAX_PER_PAGE"], fields=search_fields)
    else:
        projects = Project.query

    projects = projects.filter_by(hub_id=current_user.hub_id)

    if current_user.role != UserRoles.admin and current_user.projects:
        projects = projects.filter(Project.id.in_(p.id for p in current_user.projects))

    if price_level is not None:
        projects = projects.filter_by(price_level=price_level)

    if last_order_date is not None:
        projects = projects.filter(
            Project.last_order_date <= datetime.now().date() - pd.DateOffset(months=last_order_date)
        )

    if current_user.role != UserRoles.admin:
        projects = projects.filter_by(enabled=True)
    if search_key:
        projects = db.paginate(projects, page=1, max_per_page=current_app.config["MAX_PER_PAGE"])
        projects.total = total
        projects.page = page
    else:
        projects = projects.order_by(Project.name)
        projects = db.paginate(projects, page=page, max_per_page=current_app.config["MAX_PER_PAGE"])

    forms = {
        "add_project": AddProjectForm(),
        "edit_project": EditProjectForm(),
        "upload_projects": UploadProjectsForm(),
    }

    return render_template(
        "main/projects/projects.html",
        projects=projects,
        forms=forms,
        show_add_project=show_add_project,
        search_key=search_key,
        fast_search=fast_search,
        price_level=price_level,
        price_levels=ProjectPriceLevel,
        last_order_date=last_order_date,
    )


@bp.route("/project/add", methods=["POST"])
@login_required
@role_forbidden([UserRoles.default, UserRoles.vendor, UserRoles.initiative])
def AddProject():
    form = AddProjectForm()
    if form.validate_on_submit():
        project_name = form.project_name.data.strip()
        uid = form.uid.data.strip() if form.uid.data is not None else None
        project = Project.query.filter_by(hub_id=current_user.hub_id, name=project_name).first()
        if project is None:
            project = Project(
                name=project_name,
                uid=uid,
                hub_id=current_user.hub_id,
                tin=form.tin.data,
                phone=form.phone.data,
                email=form.email.data,
                contact=form.contact.data,
                legal_address=form.legal_address.data,
                shipping_address=form.shipping_address.data,
                note=form.note.data,
                price_level=form.price_level.data or ProjectPriceLevel.online_store,
                discount=form.discount.data or 0.0,
            )
            db.session.add(project)
            db.session.commit()
            flash(f"Клиент {project_name} добавлен.")
        else:
            flash(f"Клиент {project_name} уже существует.")
    else:
        flash_errors(form)
    return redirect(url_for("main.ShowProjects"))


@bp.route("/project/remove/<int:project_id>")
@login_required
@role_forbidden([UserRoles.default, UserRoles.vendor, UserRoles.initiative])
def RemoveProject(project_id):
    project = Project.query.filter_by(id=project_id).first()
    if project is not None:
        db.session.delete(project)
        db.session.commit()
        flash(f"Клиент {project.name} удален.")
    else:
        flash("Такого клиента не существует.")
    return redirect(url_for("main.ShowProjects"))


@bp.route("/project/edit/", methods=["POST"])
@login_required
@role_forbidden([UserRoles.default, UserRoles.vendor, UserRoles.initiative])
def EditProject():
    form = EditProjectForm()
    if form.validate_on_submit():
        project = Project.query.filter_by(id=form.project_id.data).first()
        if project is not None:
            project_name = form.project_name.data.strip()
            project.name = project_name
            project.uid = form.uid.data.strip() if form.uid.data is not None else None
            project.enabled = form.enabled.data
            project.tin = form.tin.data
            project.phone = form.phone.data
            project.email = form.email.data
            project.contact = form.contact.data
            project.legal_address = form.legal_address.data
            project.shipping_address = form.shipping_address.data
            project.note = form.note.data
            project.price_level = form.price_level.data or ProjectPriceLevel.online_store
            project.discount = form.discount.data or 0.0
            db.session.commit()
            flash(f"Клиент {project_name} изменён.")
        else:
            flash("Такого клиента не существует.")
    else:
        flash_errors(form)
    return redirect(url_for("main.ShowProjects"))


@bp.route("/projects/upload", methods=["POST"])
@login_required
@role_forbidden([UserRoles.default, UserRoles.vendor, UserRoles.initiative])
def UploadProjects():
    form = UploadProjectsForm()
    if form.validate_on_submit():
        try:
            projects, projects_order_history = projects_excel_to_df(form.projects.data)
        except Exception as exc:
            flash("Не удалось загрузить файл. Проверьте формат.", category="error")
            current_app.logger.error(exc)
            return redirect(url_for("main.ShowProjects"))
        original_projects = pd.read_sql(
            "SELECT id, uid FROM project WHERE hub_id = ?", db.engine, params=(current_user.hub_id,)
        )
        projects = projects.merge(original_projects, on="uid", how="left")

        projects["hub_id"] = current_user.hub_id
        existing_projects = projects[projects["id"].notnull()]
        new_projects = projects[projects["id"].isnull()].drop("id", axis=1)

        id_list = existing_projects["id"].to_list()

        Project.query.filter(Project.id.in_(id_list)).delete()
        db.session.commit()
        existing_projects.to_sql("project", db.engine, if_exists="append", index=False)
        new_projects.to_sql("project", db.engine, if_exists="append", index=False)

        # Clean up to save memory
        del projects
        del existing_projects
        del original_projects
        del new_projects
        del id_list

        original_projects = pd.read_sql(
            "SELECT id, uid FROM project WHERE hub_id = ?", db.engine, params=(current_user.hub_id,)
        )
        projects_order_history = projects_order_history.merge(original_projects, on="uid", how="right")
        projects_order_history.rename(columns={"id": "project_id"}, inplace=True)
        projects_order_history.drop("uid", axis=1, inplace=True)
        projects_order_history.dropna(inplace=True)

        id_list = projects_order_history["project_id"].to_list()
        ProjectOrderHistory.query.filter(ProjectOrderHistory.project_id.in_(id_list)).delete()
        db.session.commit()
        projects_order_history.to_sql("project_order_history", db.engine, if_exists="append", index=False)

        flash("Список клиентов успешно загружен.")
    else:
        flash_errors(form)
    return redirect(url_for("main.ShowProjects"))
