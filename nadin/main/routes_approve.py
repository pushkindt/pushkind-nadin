import io
import os
from copy import copy
from datetime import date, datetime, timedelta, timezone

from flask import Response, flash, redirect, render_template, request, url_for
from flask_login import current_user, login_required
from openpyxl import load_workbook
from sqlalchemy import or_
from sqlalchemy.orm.attributes import flag_modified

from nadin.extensions import db
from nadin.main.forms import ChangeQuantityForm, InitiativeForm, LeaveCommentForm, OrderApprovalForm, SplitOrderForm
from nadin.main.routes import bp
from nadin.models.hub import AppSettings, User, UserRoles, Vendor
from nadin.models.order import EventType, Order, OrderApproval, OrderEvent, OrderStatus
from nadin.models.product import Category, Product
from nadin.models.project import Project
from nadin.utils import SendEmail1C, SendEmailNotification, flash_errors, role_forbidden, role_required

################################################################################
# Approve page
################################################################################


@bp.app_template_filter()
def intersect(a, b):
    return set(a).intersection(set(b))


def get_order(order_id):
    order = Order.get_by_access(current_user)
    order = order.filter(Order.id == order_id).first()
    return order


@bp.route("/orders/<int:order_id>")
@login_required
@role_forbidden([UserRoles.default])
def show_order(order_id):

    order = get_order(order_id)
    if order is None:
        flash("Заявка с таким номером не найдена.")
        return redirect(url_for("main.ShowIndex"))

    approval_form = OrderApprovalForm()
    quantity_form = ChangeQuantityForm()
    comment_form = LeaveCommentForm()
    initiative_form = InitiativeForm()

    comment_form.notify_reviewers.choices = [(r.id, r.name) for r in order.reviewers]

    if order.project:
        initiative_form.project.choices = [(order.project["id"], order.project["name"])]
        initiative_form.project.default = order.project["id"]
        initiative_form.process()

    split_form = SplitOrderForm()

    return render_template(
        "main/approve/approve.html",
        order=order,
        comment_form=comment_form,
        approval_form=approval_form,
        quantity_form=quantity_form,
        initiative_form=initiative_form,
        split_form=split_form,
    )


@bp.route("/orders/split/<int:order_id>", methods=["POST"])
@login_required
@role_required([UserRoles.admin, UserRoles.initiative, UserRoles.purchaser])
def split_order(order_id):

    order = get_order(order_id)
    if order is None:
        flash("Заявка с таким номером не найдена.")
        return redirect(url_for("main.ShowIndex"))

    if len(order.children) > 0:
        flash("Нельзя разделять заявки, которые были объединены или разделены.")
        return redirect(url_for("main.ShowIndex"))

    if order.status != OrderStatus.new:
        flash("Нельзя модифицировать согласованную или аннулированную заявку.")
        return redirect(url_for("main.show_order", order_id=order_id))

    form = SplitOrderForm()
    if form.validate_on_submit():
        product_ids = form.products.data
        if not isinstance(product_ids, list) or len(product_ids) == 0:
            flash("Некорректный список позиции.")
            return redirect(url_for("main.show_order", order_id=order_id))

        product_lists = [[], []]

        for product in order.products:
            if str(product["id"]) in product_ids:
                product_lists[0].append(product)
            else:
                product_lists[1].append(product)

        if len(product_lists[0]) == 0 or len(product_lists[1]) == 0:
            flash("Некорректный список позиции.")
            return redirect(url_for("main.show_order", order_id=order_id))

        message_flash = "заявка разделена на заявки"

        for product_list in product_lists:
            new_order_number = Order.new_order_number(current_user.hub_id)
            message_flash += f" {new_order_number}"
            new_order = Order(number=new_order_number)
            db.session.add(new_order)
            new_order.initiative_id = order.initiative_id
            new_order.initiative = order.initiative
            now = datetime.now(tz=timezone.utc)
            new_order.products = product_list
            new_order.total = sum(product["quantity"] * product["price"] for product in new_order.products)
            new_order.project_id = order.project_id
            new_order.project = order.project
            new_order.status = OrderStatus.new
            new_order.create_timestamp = int(now.timestamp())
            new_order.hub_id = order.hub_id
            categories = [product.get("categoryId", -1) for product in new_order.products]
            new_order.categories = Category.query.filter(
                Category.id.in_(categories), Category.hub_id == current_user.hub_id
            ).all()
            vendors = [product.get("vendor") for product in new_order.products]
            new_order.vendors = Vendor.query.filter(
                Vendor.name.in_(vendors), Vendor.hub_id == current_user.hub_id
            ).all()
            new_order.parents = [order]
            message = f"заявка получена разделением из заявки {order.number}"
            event = OrderEvent(
                user_id=current_user.id,
                order_id=new_order.id,
                type=EventType.splitted,
                data=message,
                timestamp=datetime.now(tz=timezone.utc),
            )
            db.session.add(event)
            db.session.commit()
            SendEmailNotification("new", new_order)

        order.total = 0.0
        event = OrderEvent(
            user_id=current_user.id,
            order_id=order_id,
            type=EventType.splitted,
            data=message_flash,
            timestamp=datetime.now(tz=timezone.utc),
        )
        db.session.add(event)
        db.session.commit()

        flash(message_flash)

    else:
        flash_errors(form)
    return redirect(url_for("main.ShowIndex"))


@bp.route("/orders/duplicate/<int:order_id>")
@login_required
@role_required([UserRoles.admin, UserRoles.initiative, UserRoles.purchaser])
def duplicate_order(order_id):
    order = get_order(order_id)
    if order is None:
        flash("Заявка с таким номером не найдена.")
        return redirect(url_for("main.ShowIndex"))

    order_number = Order.new_order_number(current_user.hub_id)
    new_order = Order(number=order_number)
    db.session.add(new_order)

    new_order.initiative = current_user.to_dict()
    new_order.initiative_id = current_user.id

    now = datetime.now(tz=timezone.utc)

    new_order.products = order.products
    new_order.total = order.total
    new_order.project_id = order.project_id
    new_order.project = order.project
    new_order.status = OrderStatus.new
    new_order.create_timestamp = int(now.timestamp())

    new_order.hub_id = current_user.hub_id
    new_order.categories = order.categories
    new_order.vendors = order.vendors

    message = f"заявка клонирована с номером {new_order.number}"
    event = OrderEvent(
        user_id=current_user.id,
        order_id=order.id,
        type=EventType.duplicated,
        data=message,
        timestamp=datetime.now(tz=timezone.utc),
    )
    db.session.add(event)
    message = f"заявка клонирована из заявки {order.number}"
    event = OrderEvent(
        user_id=current_user.id,
        order_id=new_order.id,
        type=EventType.duplicated,
        data=message,
        timestamp=datetime.now(tz=timezone.utc),
    )
    db.session.add(event)

    db.session.commit()

    flash(f"Заявка успешно клонирована. Номер новой заявки {new_order.number}. " "Вы перемещены в новую заявку.")

    SendEmailNotification("new", new_order)

    return redirect(url_for("main.show_order", order_id=new_order.id))


@bp.route("/orders/quantity/<int:order_id>", methods=["POST"])
@login_required
@role_required([UserRoles.admin, UserRoles.initiative, UserRoles.purchaser])
def save_quantity(order_id):

    order = get_order(order_id)
    if order is None:
        flash("Заявка с таким номером не найдена.")
        return redirect(url_for("main.ShowIndex"))

    if order.status != OrderStatus.new:
        flash("Нельзя модифицировать согласованную или аннулированную заявку.")
        return redirect(url_for("main.show_order", order_id=order_id))

    form = ChangeQuantityForm()
    if form.validate_on_submit():
        product = {}
        for idx, product in enumerate(order.products):
            if form.product_id.data == product["id"]:
                break
        else:
            idx = None
            product = (
                Product.query.filter_by(id=form.product_id.data)
                .join(Vendor, Product.vendor_id == Vendor.id)
                .filter(or_(Vendor.hub_id == current_user.hub_id, Product.vendor_id == current_user.hub_id))
                .first()
            )
            if not product:
                flash("Указанный товар не найден.")
                return redirect(url_for("main.show_order", order_id=order_id))
            product = product.to_dict(current_user.price_level, current_user.discount)
            product["categoryId"] = product["cat_id"]
            product["imageUrl"] = product["image"]
            product["quantity"] = 0
            product["selectedOptions"] = [{"name": "Единицы", "value": product["measurement"]}]
            order.products.append(product)

        if product["quantity"] != form.product_quantity.data:
            message = f"{product['sku']} количество было {product['quantity']} " f"стало {form.product_quantity.data}"
            product["quantity"] = form.product_quantity.data
            event = OrderEvent(
                user_id=current_user.id,
                order_id=order_id,
                type=EventType.quantity,
                data=message,
                timestamp=datetime.now(tz=timezone.utc),
            )
            db.session.add(event)

        approvals = (
            OrderApproval.query.join(User)
            .filter(OrderApproval.order_id == order_id, User.hub_id == current_user.hub_id)
            .all()
        )
        for approval in approvals:
            db.session.delete(approval)

        order.total = sum(p["quantity"] * p["price"] for p in order.products)
        order.status = OrderStatus.new

        if idx in range(len(order.products)) and form.product_quantity.data == 0:
            order.products.pop(idx)
        flag_modified(order, "products")

        db.session.commit()

        flash(f"Позиция {product['sku']} была изменена.")

    else:
        flash_errors(form)
    return redirect(url_for("main.show_order", order_id=order_id))


@bp.route("/orders/excel1/<int:order_id>")
@login_required
@role_forbidden([UserRoles.default])
def GetExcelReport1(order_id):
    order = get_order(order_id)
    if order is None:
        flash("Заявка с таким номером не найдена.")
        return redirect(url_for("main.ShowIndex"))

    order_products = [p for p in order.products if p["quantity"] > 0]

    data_len = len(order_products)
    starting_row = 11
    wb = load_workbook(filename=os.path.join("app", "static", "upload", "template.xlsx"))
    ws = wb.active
    ws["P17"] = order.initiative.name
    if data_len > 1:
        for merged_cell in ws.merged_cells.ranges:
            if merged_cell.bounds[1] >= starting_row:
                merged_cell.shift(0, data_len)
        ws.insert_rows(starting_row, data_len - 1)
    for k, i in enumerate(range(starting_row, starting_row + data_len)):
        product = order_products[k]
        ws.row_dimensions[i].height = 50
        if data_len > 1:
            for j in range(1, 20):
                target_cell = ws.cell(row=i, column=j)
                source_cell = ws.cell(row=starting_row + data_len - 1, column=j)
                target_cell._style = copy(source_cell._style)
                target_cell.font = copy(source_cell.font)
                target_cell.border = copy(source_cell.border)
                target_cell.fill = copy(source_cell.fill)
                target_cell.number_format = copy(source_cell.number_format)
                target_cell.protection = copy(source_cell.protection)
                target_cell.alignment = copy(source_cell.alignment)
        ws.cell(i, 1).value = k + 1
        ws.cell(i, 5).value = product["name"]
        ws.cell(i, 3).value = order.project.name if order.project is not None else ""
        ws.cell(i, 7).value = product.get("vendor", "")
        ws.cell(i, 8).value = product["quantity"]
        c1 = ws.cell(i, 8).coordinate
        ws.cell(i, 10).value = product["price"]
        c2 = ws.cell(i, 10).coordinate
        ws.cell(i, 12).value = f"={c1}*{c2}"

    ws = wb["электронное согласование"]
    i = 3
    ws.cell(i, 3).value = order.number
    i += 1
    ws.cell(i, 3).value = datetime.fromtimestamp(
        order.create_timestamp, tz=timezone(timedelta(hours=3), name="Europe/Moscow")
    ).strftime("%Y-%m-%d")
    if order.status == OrderStatus.approved:
        i += 1
        for i, approval in enumerate(order.approvals, start=i):
            if approval.approved is not True:
                continue
            ws.cell(i, 3).value = approval.timestamp.astimezone(
                timezone(timedelta(hours=3), name="Europe/Moscow")
            ).strftime("%Y-%m-%d")
            ws.cell(i, 5).value = approval.user.name

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return Response(
        buffer,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment;filename=report.xlsx"},
    )


@bp.route("/orders/excel2/<int:order_id>")
@login_required
@role_forbidden([UserRoles.default])
def GetExcelReport2(order_id):
    order = get_order(order_id)
    if order is None:
        flash("Заявка с таким номером не найдена.")
        return redirect(url_for("main.ShowIndex"))

    order_products = [p for p in order.products if p["quantity"] > 0]

    starting_row = 2
    wb = load_workbook(filename=os.path.join("app", "static", "upload", "template2.xlsx"))
    ws = wb.active

    ws.title = order.project.name if order.project is not None else "Клиент не указан"

    i = starting_row
    for product in order_products:
        ws.cell(i, 1).value = product["sku"]
        ws.cell(i, 2).value = product["name"]
        if "selectedOptions" in product:
            ws.cell(i, 3).value = product["selectedOptions"][0]["value"]
        ws.cell(i, 4).value = product["price"]
        ws.cell(i, 5).value = product["quantity"]
        ws.cell(i, 6).value = product["price"] * product["quantity"]
        i += 1

    i = 3
    ws.cell(i, 9).value = order.number
    i += 1
    ws.cell(i, 9).value = datetime.fromtimestamp(
        order.create_timestamp, tz=timezone(timedelta(hours=3), name="Europe/Moscow")
    ).strftime("%Y-%m-%d")
    if order.status == OrderStatus.approved:
        i += 1
        for i, approval in enumerate(order.approvals, start=i):
            if approval.approved is not True:
                continue
            ws.cell(i, 9).value = approval.timestamp.astimezone(
                timezone(timedelta(hours=3), name="Europe/Moscow")
            ).strftime("%Y-%m-%d")
            ws.cell(i, 11).value = approval.user.name

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return Response(
        buffer,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment;filename=report.xlsx"},
    )


@bp.route("/orders/dealdone/<int:order_id>", methods=["POST"])
@login_required
@role_required([UserRoles.admin, UserRoles.purchaser])
def SetDealDone(order_id):
    order = get_order(order_id)
    if order is None:
        flash("Заявка с таким номером не найдена.")
        return redirect(url_for("main.ShowIndex"))

    if order.status == OrderStatus.cancelled:
        flash("Нельзя модифицировать аннулированную заявку.")
        return redirect(url_for("main.show_order", order_id=order_id))

    form = LeaveCommentForm()
    form.notify_reviewers.choices = [(r.id, r.name) for r in order.reviewers]
    if form.validate_on_submit():
        order.dealdone = True
        form.comment_and_send_email(order, EventType.dealdone)
        flash("Заявка законтрактована.")
        db.session.commit()
    else:
        flash_errors(form)
    return redirect(url_for("main.show_order", order_id=order_id))


def Prepare1CReport(order, excel_date):

    order_products = [p for p in order.products if p["quantity"] > 0]

    data_len = len(order_products)
    if data_len > 0:
        categories = Category.query.filter(Category.hub_id == order.initiative.hub_id).all()
        starting_row = 3
        wb = load_workbook(filename=os.path.join("app", "static", "upload", "template1C.xlsx"))
        wb.iso_dates = True
        ws = wb["Заявка"]
        for merged_cell in ws.merged_cells.ranges:
            if merged_cell.bounds[1] >= starting_row:
                merged_cell.shift(0, data_len)

        ws.insert_rows(starting_row, data_len)
        for k, i in enumerate(range(starting_row, starting_row + data_len)):
            product = order_products[k]
            ws.row_dimensions[i].height = 50

            for j in range(1, 32):
                target_cell = ws.cell(row=i, column=j)
                source_cell = ws.cell(row=starting_row + data_len, column=j)
                target_cell._style = copy(source_cell._style)
                target_cell.font = copy(source_cell.font)
                target_cell.border = copy(source_cell.border)
                target_cell.fill = copy(source_cell.fill)
                target_cell.number_format = copy(source_cell.number_format)
                target_cell.protection = copy(source_cell.protection)
                target_cell.alignment = copy(source_cell.alignment)

            # Object

            ws.cell(i, 2).value = order.project.name if order.project is not None else ""
            # Initiative

            ws.cell(i, 5).value = order.initiative.name
            ws.cell(i, 6).value = "Для собственных нужд"

            product_cat = product.get("categoryId", 0)
            for cat in categories:
                if product_cat == cat.id:
                    ws.cell(i, 10).value = cat.functional_budget
                    ws.cell(i, 24).value = cat.responsible
                    break
            else:
                ws.cell(i, 10).value = ""
                ws.cell(i, 24).value = ""

            ws.cell(i, 15).value = "Непроектные МТР и СИЗ"

            # Measurement
            if "selectedOptions" in product:
                ws.cell(i, 19).value = product["selectedOptions"][0]["value"]
                ws.cell(i, 23).value = ", ".join(p["value"] for p in product["selectedOptions"][1:])
            # Product Name
            ws.cell(i, 20).value = product.get("name", "")
            # Quantity
            ws.cell(i, 22).value = product.get("quantity", "")
            ws.cell(i, 29).value = excel_date

            ws.cell(i, 31).value = product.get("price", "")

            ws.cell(i, 30).value = product.get("vendor", "")

        i += 6
        ws.cell(i, 20).value = order.number
        i += 1
        ws.cell(i, 20).value = datetime.fromtimestamp(
            order.create_timestamp,
            tz=timezone(timedelta(hours=3), name="Europe/Moscow"),
        ).strftime("%Y-%m-%d")
        if order.status == OrderStatus.approved:
            i += 1
            for i, approval in enumerate(order.approvals, start=i):
                if approval.approved is not True:
                    continue
                ws.cell(i, 20).value = approval.timestamp.astimezone(
                    timezone(timedelta(hours=3), name="Europe/Moscow")
                ).strftime("%Y-%m-%d")
                ws.cell(i, 22).value = approval.user.name
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
    return None


@bp.route("/orders/excel1C/<int:order_id>")
@login_required
@role_forbidden([UserRoles.default])
def GetExcelReport1C(order_id):
    order = get_order(order_id)
    if order is None:
        flash("Заявка с таким номером не найдена.")
        return redirect(url_for("main.ShowIndex"))
    try:
        excel_date = request.args.get(
            "date",
            default=date.today(),
            type=lambda x: datetime.strptime(x, "%Y-%m-%d").date(),
        )
    except ValueError:
        excel_date = date.today()
    excel_send = request.args.get("send", default=False, type=bool)
    data = Prepare1CReport(order, excel_date)
    if data is None:
        flash("Не удалось получить выгрузку.")
        return redirect(url_for("main.show_order", order_id=order_id))
    if excel_send is False:
        return Response(
            data,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment;filename=pushkind_{order.number}.xlsx"},
        )

    app_data = AppSettings.query.filter_by(hub_id=current_user.hub_id).first()
    if app_data is not None and app_data.email_1C is not None:
        SendEmail1C([app_data.email_1C], order, data.read())
        message = f"отправлена на {app_data.email_1C}"
        event = OrderEvent(
            user_id=current_user.id,
            order_id=order.id,
            type=EventType.exported,
            data=message,
            timestamp=datetime.now(tz=timezone.utc),
        )
        db.session.add(event)
        order.exported = True
        db.session.commit()
        flash(f"Заявка отправлена на {app_data.email_1C}")
    else:
        flash("Email для отправки в 1С не настроен администратором.")
    return redirect(url_for("main.show_order", order_id=order_id))


@bp.route("/orders/approval/<int:order_id>", methods=["POST"])
@login_required
@role_required([UserRoles.validator])
def SaveApproval(order_id):
    order = get_order(order_id)
    if order is None:
        flash("Заявка с таким номером не найдена.")
        return redirect(url_for("main.ShowIndex"))

    if order.status == OrderStatus.cancelled:
        flash("Нельзя модифицировать аннулированную заявку.")
        return redirect(url_for("main.show_order", order_id=order_id))

    form = OrderApprovalForm()
    if form.validate_on_submit():

        message = form.comment.data.strip()
        if len(message) == 0:
            message = None

        user_approval = OrderApproval.query.filter_by(
            user_id=current_user.id,
            order_id=order.id,
            product_id=form.product_id.data,
            remark=message,
        ).first()

        if user_approval is not None:
            flash("Вы уже выполнили это действие.")
            return redirect(url_for("main.show_order", order_id=order_id))

        last_status = order.status

        if form.product_id.data is None:
            OrderApproval.query.filter_by(order_id=order_id, user_id=current_user.id).delete()

            position_disapprovals = OrderApproval.query.filter_by(order_id=order_id)
            position_disapprovals = position_disapprovals.join(User)

            for disapproval in position_disapprovals:
                db.session.delete(disapproval)

            order_approval = OrderApproval(
                order_id=order_id,
                product_id=None,
                user_id=current_user.id,
                remark=message,
            )
            db.session.add(order_approval)
            event = OrderEvent(
                user_id=current_user.id,
                order_id=order_id,
                type=EventType.approved,
                data=message,
                timestamp=datetime.now(tz=timezone.utc),
            )
        else:
            OrderApproval.query.filter_by(order_id=order_id, user_id=current_user.id, product_id=None).delete()
            if form.product_id.data == 0:
                event = OrderEvent(
                    user_id=current_user.id,
                    order_id=order_id,
                    type=EventType.disapproved,
                    data=message,
                    timestamp=datetime.now(tz=timezone.utc),
                )
                product_approval = OrderApproval(
                    order_id=order_id,
                    product_id=0,
                    user_id=current_user.id,
                    remark=message,
                )
                db.session.add(product_approval)
            else:
                product = {}
                for product in order.products:
                    if form.product_id.data == product["id"]:
                        break
                else:
                    flash("Указанный позиция не найдена в заявке.")
                    return redirect(url_for("main.show_order", order_id=order_id))
                product_approval = OrderApproval.query.filter_by(
                    order_id=order_id,
                    user_id=current_user.id,
                    product_id=form.product_id.data,
                ).first()
                if product_approval is None:
                    product_approval = OrderApproval(
                        order_id=order_id,
                        product_id=form.product_id.data,
                        user_id=current_user.id,
                    )
                    db.session.add(product_approval)
                product_approval.remark = message
                message = f"к позиции \"{product['name']}\" {message or ''}"
                event = OrderEvent(
                    user_id=current_user.id,
                    order_id=order_id,
                    type=EventType.disapproved,
                    data=message,
                    timestamp=datetime.now(tz=timezone.utc),
                )
        db.session.add(event)
        order.update_status()
        db.session.commit()
        flash("Согласование сохранено.")

        if order.status != last_status:
            if order.status == OrderStatus.fulfilled:
                SendEmailNotification("approved", order)
                app_data = AppSettings.query.filter_by(hub_id=current_user.hub_id).first()
                if app_data is not None and app_data.email_1C is not None and app_data.notify_1C is True:
                    data = Prepare1CReport(order, date.today() + timedelta(days=14))
                    if data is not None:
                        SendEmail1C([app_data.email_1C], order, data.read())

            elif order.status == OrderStatus.cancelled:
                SendEmailNotification("disapproved", order)

    else:
        flash_errors(form)
    return redirect(url_for("main.show_order", order_id=order_id))


@bp.route("/orders/parameters/<int:order_id>", methods=["POST"])
@login_required
@role_required([UserRoles.admin, UserRoles.initiative, UserRoles.validator, UserRoles.purchaser])
def SaveParameters(order_id):
    order = get_order(order_id)
    if order is None:
        flash("Заявка с таким номером не найдена.")
        return redirect(url_for("main.ShowIndex"))

    if order.status != OrderStatus.new:
        flash("Нельзя модифицировать согласованную или аннулированную заявку.")
        return redirect(url_for("main.show_order", order_id=order_id))

    form = InitiativeForm()

    projects = Project.query.filter(Project.hub_id == current_user.hub_id).order_by(Project.name).all()
    categories = Category.query.filter(Category.hub_id == current_user.hub_id).all()

    form.categories.choices = [(c.id, c.name) for c in categories]
    form.project.choices = [(p.id, p.name) for p in projects]

    if form.validate_on_submit() is True:
        new_project = Project.query.filter_by(id=form.project.data, hub_id=current_user.hub_id).first()
        if new_project is not None and (order.project is None or order.project_id != new_project.id):
            message = f'Клиент изменён «{order.project.name if order.project else ""}» на «{new_project.name}»'
            event = OrderEvent(
                user_id=current_user.id,
                order_id=order_id,
                type=EventType.project,
                data=message,
                timestamp=datetime.now(tz=timezone.utc),
            )
            db.session.add(event)
            order.project = new_project

        order.categories = Category.query.filter(
            Category.id.in_(form.categories.data),
            Category.hub_id == current_user.hub_id,
        ).all()
        OrderApproval.query.filter_by(order_id=order.id).delete()
        db.session.commit()

        flash("Параметры заявки успешно сохранены.")
    else:
        flash_errors(form)
    return redirect(url_for("main.show_order", order_id=order_id))


@bp.route("/orders/comment/<int:order_id>", methods=["POST"])
@login_required
@role_required([UserRoles.admin, UserRoles.initiative, UserRoles.validator, UserRoles.purchaser])
def LeaveComment(order_id):
    order = get_order(order_id)
    if order is None:
        flash("Заявка с таким номером не найдена.")
        return redirect(url_for("main.ShowIndex"))
    form = LeaveCommentForm()
    form.notify_reviewers.choices = [(r.id, r.name) for r in order.reviewers]
    if form.validate_on_submit():
        form.comment_and_send_email(order, EventType.commented)
        db.session.commit()
    else:
        flash_errors(form)
    return redirect(url_for("main.show_order", order_id=order_id))


@bp.route("/orders/process/<int:order_id>")
@login_required
@role_required([UserRoles.admin, UserRoles.purchaser])
def ProcessHubOrder(order_id):
    order = get_order(order_id)
    if order is None:
        flash("Заявка с таким номером не найдена.")
        return redirect(url_for("main.ShowIndex"))

    if order.status == OrderStatus.cancelled:
        flash("Нельзя отправить поставщику аннулированную заявку.")
        return redirect(url_for("main.show_order", order_id=order_id))

    message = "Заявка была отправлена поставщикам: "
    message += ", ".join(vendor.name for vendor in order.vendors)

    event = OrderEvent(
        user_id=current_user.id,
        order_id=order_id,
        type=EventType.purchased,
        data=message,
        timestamp=datetime.now(tz=timezone.utc),
    )
    db.session.add(event)
    order.purchased = True
    db.session.commit()

    flash(message)

    return redirect(url_for("main.show_order", order_id=order_id))


@bp.route("/orders/cancel/<int:order_id>", methods=["POST"])
@login_required
@role_required([UserRoles.admin, UserRoles.initiative])
def CancelOrder(order_id):
    order = get_order(order_id)
    if order is None:
        flash("Заявка с таким номером не найдена.")
        return redirect(url_for("main.ShowIndex"))

    if order.status == OrderStatus.cancelled:
        flash("Нельзя аннулировать аннулированную заявку.")
        return redirect(url_for("main.show_order", order_id=order_id))

    form = LeaveCommentForm()
    form.notify_reviewers.choices = [(r.id, r.name) for r in order.reviewers]
    if form.validate_on_submit():
        order.status = OrderStatus.cancelled
        order.total = 0
        form.comment_and_send_email(order, EventType.cancelled)
        db.session.commit()
        flash("Заявка аннулирована.")
    else:
        flash_errors(form)
    return redirect(url_for("main.show_order", order_id=order_id))
